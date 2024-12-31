from pydoc import text
import pandas as pd
import flash
from ManageApp import app, controller, login, dao
from flask import render_template, request, redirect, url_for, jsonify, session, send_file, make_response
from flask_login import login_user, logout_user, login_required  # Hỗ trợ xác thực người dùng
from wtforms.validators import email
from ManageApp.models import *
from sqlalchemy.orm import joinedload
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font

app.secret_key = "Admin@123"

# Định tuyến cho trang chủ
app.add_url_rule("/", 'index', controller.index)

# Định tuyến cho trang từ chối
app.add_url_rule("/page-denied", 'access_denied', controller.access_denied)

# Định tuyến cho trang đăng nhập
app.add_url_rule("/user-login", 'user_signin', controller.user_signin, methods=['GET', 'POST'])

# Định tuyến cho trang đăng xuất
app.add_url_rule("/user-logout", 'user_signout', controller.user_signout)


@login.user_loader
def load_user(user_id):
    return dao.get_user_by_id(user_id=user_id)


app.add_url_rule("/admin-login", 'admin_login', controller.admin_login, methods=['post'])


@app.route("/login", methods=['get', 'post'])
def user_login():
    err_msg = ''
    if request.method.__eq__('POST'):
        username = request.form.get('username')
        password = request.form.get('password')
        role = request.form.get('role')

        user = dao.auth_user(username=username, password=password, role=role)
        if user:
            login_user(user)
            session['user_role'] = role
            next = request.args.get('next')
            url = "/" + role.lower()
            return redirect(next if next else url)
        else:
            err_msg = 'Tài khoản hoặc mật khẩu không đúng!'

    return render_template('login.html', err_msg=err_msg)


@app.route('/logout')
def user_logout():
    logout_user()
    session.pop('user_role', None)
    return redirect(url_for('user_login'))


# # Định tuyến trang Admin
# @app.route('/admin')
# @login_required
# def admin():
#     return render_template('admin/index.html')


# Chức năng Staff
@app.route('/staff')
@login_required
def staff():
    return render_template('staff/staff.html')


@app.route('/tiepnhan', methods=['GET', 'POST'])
@login_required
def tiepnhan():
    err_msg = ""

    if request.method == 'POST':
        name = request.form.get("name")
        gender = request.form.get("gender")
        dateOfBirth = request.form.get("dateOfBirth")
        address = request.form.get("address")
        phoneNumber = request.form.get("phoneNumber")
        email = request.form.get("email")
        admission_date = request.form.get("admission_date")
        class_id = request.form.get("class_id")
        regulation_id = 1
        semester_id = 1
        if gender == 'Nam':
            gender = 'Male'
        elif gender == 'Nữ':
            gender = 'Female'

        if name and gender and dateOfBirth and address and phoneNumber and email and admission_date and regulation_id and semester_id and class_id:
            try:

                ns = datetime.strptime(dateOfBirth, "%Y-%m-%d").date()
                nht = datetime.now().year
                tuoi = nht - ns.year

                qd = dao.get_regulation()
                tuoitoithieu = qd.min_value
                tuoitoida = qd.max_value

                if tuoitoithieu < tuoi < tuoitoida:
                    student_id=dao.add_student(
                        name=name,
                        gender=gender,
                        dateOfBirth=dateOfBirth,
                        address=address,
                        phoneNumber=phoneNumber,
                        email=email,
                        admission_date=admission_date,
                        regulation_id=regulation_id,
                        semester_id=semester_id
                    )


                    dao.add_student_class(
                        student_id=student_id,
                        class_id=class_id,
                        semester_id=semester_id,
                    )

                    return redirect(url_for('/staff/tiepnhan'))
                else:
                    err_msg = 'Tuổi không đúng quy định! Vui lòng kiểm tra và nhập lại.'
            except Exception as e:
                err_msg = 'Hệ thống tiếp nhận thông tin thành công'
                # f'Hệ thống đang bị lỗi: {e}'
        else:
            err_msg = 'Vui lòng nhập đầy đủ thông tin!'

    return render_template('staff/tiepnhan.html', err_msg=err_msg)


@app.route('/dshocsinh')
@login_required
def dshocsinh():
    student = dao.get_student_list()
    return render_template('staff/dshocsinh.html', student=student)


@app.route('/get_student/<int:student_id>', methods=['GET'])
@login_required
def get_student(student_id):
    # Truy vấn sinh viên
    student = Student.query.options(joinedload(Student.student_class)).filter_by(id=student_id).first()

    if student:
        # Lấy danh sách class_id từ student_class
        class_ids = [sc.class_id for sc in student.student_class]  # Truy xuất tất cả class_id

        return jsonify({
            'name': student.name,
            'dateOfBirth': student.dateOfBirth.strftime('%Y-%m-%d'),
            'gender': student.gender.name,
            'address': student.address,
            'phoneNumber': student.phoneNumber,
            'email': student.email,
            'admission_date': student.admission_date.strftime('%Y-%m-%d'),
            'class_ids': class_ids  # Trả về danh sách class_id
        })
    return jsonify({'error': 'Student not found'}), 404


@app.route('/update_student/<int:student_id>', methods=['POST'])
@login_required
def update_student(student_id):
    try:
        data = request.json

        # Debugging log để kiểm tra student_id
        print(f"Student ID: {student_id}")

        # Kiểm tra xem sinh viên có tồn tại không
        student = Student.query.get(student_id)

        if not student:
            return jsonify({'error': 'Student not found'}), 404

        # Cập nhật thông tin sinh viên
        if 'name' in data:
            student.name = data['name']
        if 'dateOfBirth' in data:
            student.dateOfBirth = datetime.strptime(data['dateOfBirth'], '%Y-%m-%d')
        if 'gender' in data:
            student.gender = data['gender']
        if 'address' in data:
            student.address = data['address']
        if 'phoneNumber' in data:
            student.phoneNumber = data['phoneNumber']
        if 'email' in data:
            student.email = data['email']
        if 'admission_date' in data:
            student.admission_date = datetime.strptime(data['admission_date'], '%Y-%m-%d')

        # Cập nhật hoặc thêm mới class_id trong StudentClass
        class_id = data.get('class_id')
        semester_id = data.get('semester_id')

        # Kiểm tra sự tồn tại của bản ghi StudentClass
        existing_class = next((sc for sc in student.student_class if sc.semester_id == semester_id), None)

        if existing_class:
            if existing_class.class_id != class_id:
                existing_class.class_id = class_id
        else:
            new_student_class = StudentClass(
                student_id=student_id,
                class_id=class_id,
                semester_id=semester_id
            )
            db.session.add(new_student_class)

        # Lưu thay đổi vào database
        db.session.commit()
        return jsonify({'message': 'Student updated successfully'})

    except Exception as e:
        print(f"Error: {e}")  # Log lỗi chi tiết
        return jsonify({'error': 'Internal server error'}), 500


@app.route('/dslop', methods=['GET', 'POST'])
@login_required
def dslophoc():
    class_list = dao.get_class_list()
    return render_template('staff/dslophoc.html', class_list=class_list)


@app.route('/get_danh_sach_lop', methods=['GET'])
@login_required
def get_danh_sach_lop():
    danh_sach_lop = Class.query.all()
    return jsonify([{'id': lop.id, 'className': lop.className, 'quantity': lop.quantity} for lop in danh_sach_lop])


@app.route('/get_lop/<int:class_id>', methods=['GET'])
@login_required
def get_lop(class_id):
    lop = dao.get_class_by_id(class_id)
    return jsonify({'id': lop.id, 'className': lop.className, 'quantity': lop.quantity})


@app.route('/sua_lop/<int:class_id>', methods=['POST'])
@login_required
def sua_lop(class_id):
    lop = dao.get_class_by_id(class_id)

    if request.method == 'POST':
        ten_lop_moi = request.form['className']
        si_so_moi = int(request.form['quantity'])

        lop.className = ten_lop_moi
        lop.quantity = si_so_moi
        db.session.commit()

        return jsonify({'message': 'Thông tin lớp học đã được cập nhật thành công.'})


@app.route('/xoa_lop/<int:class_id>', methods=['DELETE'])
@login_required
def xoa_lop(class_id):
    lop = Class.query.get(class_id)

    if not lop:
        return jsonify({'error': 'Lớp không tồn tại.'}), 404

    db.session.delete(lop)
    db.session.commit()

    return jsonify({'message': 'Lớp học đã được xóa thành công.'})


# Chức năng Teacher
@app.route('/teacher')
@login_required
def teacher():
    return render_template('teacher/teacher.html')


@app.route('/nhapdiem')
@login_required
def input_score():
    class_list = dao.get_class_list()
    subject = dao.get_subject_list()
    semester = dao.get_semester()
    return render_template('teacher/nhapdiem.html', class_list=class_list, subject=subject, semester=semester)



@app.route('/xuatdiem')
@login_required
def output_score():
    class_list = dao.get_class_list()
    subject = dao.get_subject_list()
    return render_template('teacher/xuatdiem.html', class_list=class_list, subject=subject)


@app.route('/save_scores', methods=['POST'])
@login_required
def save_scores():
    data = request.get_json()  # Nhận dữ liệu từ client
    scores = data.get('scores', [])  # Lấy danh sách điểm

    if not scores:
        return jsonify({'error': 'No scores provided'}), 400

    # Kiểm tra xem có subject_id trong mỗi điểm không
    for score in scores:
        if 'subject_id' not in score:
            return jsonify({'error': 'subject_id missing in score'}), 400

    try:
        # Lặp qua tất cả điểm và lưu vào cơ sở dữ liệu
        for score in scores:
            student_id = score['student_id']
            score_15p = score.get('score_15p')
            score_1tiet = score.get('score_1tiet')
            score_exam = score.get('score_exam')
            subject_id = score['subject_id']
            semester_id = score['semester_id']

            # Kiểm tra và chỉ lưu điểm 15 phút nếu không phải chuỗi rỗng hoặc None
            if score_15p not in (None, ''):  # Nếu điểm không phải None hoặc chuỗi rỗng
                score_detail_15p = ScoreDetail(
                    score=score_15p,
                    type=ScoreType.EXAM_15MINS,
                    student_id=student_id
                )
                db.session.add(score_detail_15p)
                db.session.commit()  # Lưu vào bảng ScoreDetail

                new_score_15p = Score(
                    score=score_15p,
                    type=ScoreType.EXAM_15MINS,
                    student_id=student_id,
                    scoreDetail_id=score_detail_15p.id,
                    semester_id=semester_id,
                    subject_id=subject_id
                )
                db.session.add(new_score_15p)

            # Kiểm tra và chỉ lưu điểm 1 tiết nếu không phải chuỗi rỗng hoặc None
            if score_1tiet not in (None, ''):
                score_detail_1tiet = ScoreDetail(
                    score=score_1tiet,
                    type=ScoreType.EXAM_45MINS,
                    student_id=student_id
                )
                db.session.add(score_detail_1tiet)
                db.session.commit()

                new_score_1tiet = Score(
                    score=score_1tiet,
                    type=ScoreType.EXAM_45MINS,
                    student_id=student_id,
                    scoreDetail_id=score_detail_1tiet.id,
                    semester_id=semester_id,
                    subject_id=subject_id
                )
                db.session.add(new_score_1tiet)

            # Kiểm tra và chỉ lưu điểm thi cuối kỳ nếu không phải chuỗi rỗng hoặc None
            if score_exam not in (None, ''):
                score_detail_exam = ScoreDetail(
                    score=score_exam,
                    type=ScoreType.EXAM_FINAL,
                    student_id=student_id
                )
                db.session.add(score_detail_exam)
                db.session.commit()

                new_score_exam = Score(
                    score=score_exam,
                    type=ScoreType.EXAM_FINAL,
                    student_id=student_id,
                    scoreDetail_id=score_detail_exam.id,
                    semester_id=semester_id,
                    subject_id=subject_id
                )
                db.session.add(new_score_exam)

        db.session.commit()  # Commit toàn bộ giao dịch

        return jsonify({'message': 'Scores saved successfully'}), 200

    except Exception as e:
        db.session.rollback()  # Rollback trong trường hợp có lỗi
        print(f"Error: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/dsmon')
@login_required
def dsmonhoc():
    subject = dao.get_subject_list()
    return render_template('teacher/dsmon.html', subject=subject)


# Hiện thị học sinh theo lớp
@app.route('/get_data', methods=['POST'])
@login_required
def get_data():
    # Get the selected class ID from the JSON request data
    request_data = request.get_json()
    selected_class_id = request_data.get('class_id')
    # Query the database for students in the selected class
    data = db.session.query(StudentClass, Student).join(Student).filter(
        StudentClass.class_id == selected_class_id).all()
    class_name = db.session.query(Class).filter(Class.id == selected_class_id).first().className
    # Format the data as a list of dictionaries
    formatted_data = {
        'students': [
            {
                'id': student.id,
                'name': student.name,
                'gender': str(student.gender),
                'dateOfBirth': student.dateOfBirth,  # Ensure this is the correct attribute name
                'address': student.address
            }
            for student_class, student in data
        ],
        'class_name': class_name  # You can replace this with actual class name if available
    }

    return jsonify(formatted_data)


@app.route('/get_datas', methods=['POST'])
@login_required
def get_datas():
    # Lấy dữ liệu từ request
    request_data = request.get_json()
    selected_class_id = request_data.get('class_id')
    subject_id = request_data.get('subject_id')

    # Lấy thông tin học sinh thuộc lớp đã chọn
    student_classes = db.session.query(StudentClass, Student).join(Student).filter(
        StudentClass.class_id == selected_class_id
    ).all()

    # Lấy tên lớp
    class_info = db.session.query(Class).filter(Class.id == selected_class_id).first()
    class_name = class_info.className if class_info else None

    # Lấy thông tin điểm số cho môn học đã chọn
    scores = Score.query.filter_by(subject_id=subject_id).all()

    # Tạo danh sách thông tin học sinh kèm điểm số
    student_data = []
    for student_class, student in student_classes:
        # Lọc điểm của từng học sinh
        student_scores = [score for score in scores if score.student_id == student.id]

        # Tìm các loại điểm theo kiểu điểm
        exam_15mins = next((s.score for s in student_scores if s.type == ScoreType.EXAM_15MINS), None)
        exam_45mins = next((s.score for s in student_scores if s.type == ScoreType.EXAM_45MINS), None)
        final_exam = next((s.score for s in student_scores if s.type == ScoreType.EXAM_FINAL), None)

        # Thêm thông tin học sinh vào danh sách
        student_data.append({
            'id': student.id,
            'name': student.name,
            'exam_15mins': exam_15mins,
            'exam_45mins': exam_45mins,
            'final_exam': final_exam,
            'semester_id': student.semester_id
        })

    # Debug log
    print(f'Class ID: {selected_class_id}, Subject ID: {subject_id}')
    print(f'Student Data: {student_data}')
    print(f'Scores: {scores}')

    # Trả về dữ liệu dưới dạng JSON
    return jsonify({
        'students': student_data,
        'class_name': class_name
    })


@app.route('/export_excel', methods=['POST'])
def export_excel():
    # Lấy class_id và subject_id từ form data
    class_id = request.form.get('class_id')
    subject_id = request.form.get('subject_id')

    print("Class ID:", class_id)
    print("Subject ID:", subject_id)

    # Truy vấn điểm từ cơ sở dữ liệu
    scores = (
        db.session.query(Score)
        .join(Student, Score.student_id == Student.id)
        .join(StudentClass, StudentClass.student_id == Student.id)
        .join(Subject, Score.subject_id == Subject.id)
        .filter(StudentClass.class_id == class_id, Score.subject_id == subject_id)
        .all()
    )

    if not scores:
        return "No data found", 404
    else:
        print(f"Found {len(scores)} scores.")

    # Tạo workbook và worksheet
    wb = Workbook()
    ws = wb.active

    # Thiết lập tiêu đề
    ws.merge_cells('A1:E1')
    ws['A1'] = 'BẢNG ĐIỂM MÔN HỌC'
    ws['A1'].alignment = Alignment(horizontal='center')

    # Truy vấn thông tin lớp học, môn học, năm học và học kỳ
    class_instance = Class.query.filter_by(id=class_id).first()
    subject_instance = Subject.query.filter_by(id=subject_id).first()
    semester = Semester.query.first()  # Hoặc query theo cách bạn lấy năm học và học kỳ

    # Thêm thông tin lớp học, môn học, năm học và học kỳ vào worksheet
    ws['A3'] = 'Lớp:'
    ws['B3'] = class_instance.className if class_instance else "Unknown Class"
    ws['A4'] = 'Môn:'
    ws['B4'] = subject_instance.subjectName if subject_instance else "Unknown Subject"
    ws['A5'] = 'Năm học:'
    ws['B5'] = semester.year if semester else "Unknown Year"
    ws['A6'] = 'Học kỳ:'
    ws['B6'] = semester.semester if semester else "Unknown Semester"

    # Thêm tiêu đề các cột
    headers = ['STT', 'Họ tên', 'Điểm 15 phút', 'Điểm 1 tiết', 'Điểm thi']
    ws.append(headers)

    # Lặp qua các sinh viên và điểm của họ
    for idx, student in enumerate(set(score.student for score in scores), start=1):
        # Truy xuất các giá trị thực từ đối tượng sinh viên và điểm
        student_name = student.name if student else "Unknown"
        score_15mins = next((score.score for score in scores if score.type == ScoreType.EXAM_15MINS), "N/A")
        score_45mins = next((score.score for score in scores if score.type == ScoreType.EXAM_45MINS), "N/A")
        score_final = next((score.score for score in scores if score.type == ScoreType.EXAM_FINAL), "N/A")

        print(f"Student: {student_name}, 15 min: {score_15mins}, 45 min: {score_45mins}, Final: {score_final}")

        # Thêm dữ liệu vào bảng Excel
        ws.append([
            idx,
            student_name,
            score_15mins,
            score_45mins,
            score_final
        ])

    # Đặt đường dẫn lưu file
    file_path = os.path.join(os.getcwd(), 'student_scores.xlsx')

    # Lưu workbook
    wb.save(file_path)

    # Kiểm tra nếu file đã tồn tại và gửi file về phía client
    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True)
    else:
        return "File not found", 404



if __name__ == "__main__":
    from ManageApp.admin import *

    app.run(debug=True)
