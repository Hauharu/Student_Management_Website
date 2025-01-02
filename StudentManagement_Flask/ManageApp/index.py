import io
import flash
from sqlalchemy import func, case
from ManageApp import app, controller, login, dao
from flask import render_template, request, redirect, url_for, jsonify, session, send_file, make_response, flash
from flask_login import login_user, logout_user, login_required  # Hỗ trợ xác thực người dùng
from wtforms.validators import email
from ManageApp.models import *
from sqlalchemy.orm import joinedload
import os
from ManageApp.dao import send_email
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font

app.secret_key = "Admin@123"

# Định tuyến cho trang chủ
app.add_url_rule("/", 'index', controller.index)

# Định tuyến cho trang từ chối
app.add_url_rule("/page-denied", 'access_denied', controller.access_denied)

# Định tuyến cho trang đăng nhập
app.add_url_rule("/user-login", 'user_signin', controller.user_signin, methods=['GET', 'POST'])

# Định tuyến cho trang đăng xuất
app.add_url_rule("/user-logout", 'user_signout', controller.user_signout)


@login.user_loader
def load_user(user_id):
    return dao.get_user_by_id(user_id=user_id)


app.add_url_rule("/admin-login", 'admin_login', controller.admin_login, methods=['post'])


# Dinh tuyen trang dang nhap
@app.route("/login", methods=['get', 'post'])
def user_login():
    err_msg = ''
    if request.method.__eq__('POST'):
        username = request.form.get('username')
        password = request.form.get('password')
        role = request.form.get('role')

        user = dao.auth_user(username=username, password=password, role=role)
        if user:
            login_user(user)
            session['user_role'] = role
            next = request.args.get('next')
            url = "/" + role.lower()
            return redirect(next if next else url)
        else:
            err_msg = 'Tài khoản hoặc mật khẩu không đúng!'

    return render_template('login.html', err_msg=err_msg)


# Dinh tuyen trang calendar
@app.route('/calendar')
def calendar():
    return render_template('calendar.html')


@app.route('/logout')
def user_logout():
    logout_user()
    session.clear()

    return redirect(url_for('user_login'))


# Dinh tuyen tang tu choi
@app.route("/change_password", methods=['GET', 'POST'])
@login_required
def change_password():
    error_msg = None
    success_msg = None
    if request.method == 'POST':
        try:
            o_password = request.form['o-password']
            n_password = request.form['n-password']
            c_password = request.form['c-password']

            hashed_o_password = hashlib.md5(o_password.encode('utf-8')).hexdigest()
            stored_password = dao.get_password_by_user_id(current_user.id)

            if hashed_o_password != stored_password:
                error_msg = "Mật khẩu cũ không đúng."
            elif n_password != c_password:
                error_msg = "Mật khẩu xác nhận không khớp."
            else:
                hashed_n_password = hashlib.md5(n_password.encode('utf-8')).hexdigest()

                dao.change_password(user_id=current_user.id, password=hashed_n_password)

                success_msg = "Thay đổi mật khẩu thành công! Đang chuyển hướng..."
                logout_user()
                session.clear()
        except Exception as ex:
            error_msg = f"Đã xảy ra lỗi: {str(ex)}"

    return render_template('change_password.html', error_msg=error_msg, success_msg=success_msg)


# # Định tuyến trang Admin
# @app.route('/admin')
# @login_required
# def admin():
#     return render_template('admin/index.html')


# Chức năng Staff
@app.route('/staff')
@login_required
def staff():
    return render_template('staff/staff.html')


from sqlalchemy.exc import IntegrityError


@app.route('/delete-student/<int:student_id>', methods=['DELETE'])
def delete_student(student_id):
    try:
        # Kiểm tra xem học sinh có tồn tại không
        student = Student.query.get(student_id)
        if not student:
            return jsonify({'success': False, 'error': 'Học sinh không tồn tại'})

        # Xóa tất cả các bản ghi liên quan trong bảng score
        scores = Score.query.join(ScoreDetail).filter(ScoreDetail.student_id == student_id).all()
        for score in scores:
            db.session.delete(score)

        # Xóa tất cả các bản ghi liên quan trong bảng score_detail
        ScoreDetail.query.filter_by(student_id=student_id).delete()

        # Xóa học sinh
        db.session.delete(student)
        db.session.commit()
        return jsonify({'success': True})
    except IntegrityError as e:
        db.session.rollback()  # Rollback nếu có lỗi khóa ngoại
        return jsonify({'success': False, 'error': 'Không thể xóa do ràng buộc khóa ngoại: ' + str(e)})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})


# Dinh tuyen trang tiep nhan
@app.route('/tiepnhan', methods=['GET', 'POST'])
@login_required
def tiepnhan():
    err_msg = ""
    success_msg = ""
    info_student = None  # Biến lưu thông tin học sinh sau khi tiếp nhận

    if request.method == 'POST':
        name = request.form.get("name")
        gender = request.form.get("gender")
        dateOfBirth = request.form.get("dateOfBirth")
        address = request.form.get("address")
        phoneNumber = request.form.get("phoneNumber")
        email = request.form.get("email")
        admission_date = request.form.get("admission_date")
        class_name = request.form.get("class_id")
        regulation_id = 1
        semester_id = 1

        if gender == 'Nam':
            gender = 'Male'
        elif gender == 'Nữ':
            gender = 'Female'

        if name and gender and dateOfBirth and address and phoneNumber and email and admission_date and class_name:
            try:
                ns = datetime.strptime(dateOfBirth, "%Y-%m-%d").date()
                tuoi = datetime.now().year - ns.year

                qd = dao.get_regulation()
                tuoitoithieu = qd.min_value
                tuoitoida = qd.max_value

                if tuoitoithieu < tuoi < tuoitoida:
                    if dao.check_duplicate(name=name, phoneNumber=phoneNumber, email=email):
                        err_msg = "Học sinh đã tồn tại! Vui lòng kiểm tra lại."
                        flash(err_msg, "error")
                        return render_template('staff/tiepnhan.html', err_msg=err_msg)

                    class_obj = dao.get_class_by_name(class_name)
                    if not class_obj:
                        err_msg = 'Lớp không tồn tại! Vui lòng kiểm tra lại.'
                        flash(err_msg, "error")
                        return render_template('staff/tiepnhan.html', err_msg=err_msg)

                    class_id = class_obj.id

                    student_id = dao.add_student(
                        name=name,
                        gender=gender,
                        dateOfBirth=dateOfBirth,
                        address=address,
                        phoneNumber=phoneNumber,
                        email=email,
                        admission_date=admission_date,
                        regulation_id=regulation_id,
                        semester_id=semester_id
                    )

                    dao.add_student_class(
                        student_id=student_id,
                        class_id=class_id,
                        semester_id=semester_id,
                    )

                    info_student = dao.get_student_by_id(student_id)

                    # Gửi email
                    send_email(to_email=email, student_name=name)

                    flash("Thêm học sinh thành công! Email đã được gửi.", "success")
                else:
                    err_msg = 'Tuổi không đúng quy định! Vui lòng kiểm tra và nhập lại.'
                    flash(err_msg, "error")
            except Exception as e:
                db.session.rollback()
                err_msg = f'Hệ thống đang bị lỗi: {str(e)}'
                flash(err_msg, "error")
        else:
            err_msg = 'Vui lòng nhập đầy đủ thông tin!'
            flash(err_msg, "error")

    return render_template('staff/tiepnhan.html', err_msg=err_msg, info_student=info_student)


# Dinh tuyen trang danh sach hoc sinh
@app.route('/dshocsinh')
@login_required
def dshocsinh():
    student = dao.get_student_list()
    return render_template('staff/dshocsinh.html', student=student)


@app.route('/get_student/<int:student_id>', methods=['GET'])
@login_required
def get_student(student_id):
    # Truy vấn sinh viên
    student = Student.query.options(joinedload(Student.student_class)).filter_by(id=student_id).first()

    if student:
        # Lấy danh sách class_id từ student_class
        class_ids = [sc.class_id for sc in student.student_class]  # Truy xuất tất cả class_id

        return jsonify({
            'name': student.name,
            'dateOfBirth': student.dateOfBirth.strftime('%Y-%m-%d'),
            'gender': student.gender.name,
            'address': student.address,
            'phoneNumber': student.phoneNumber,
            'email': student.email,
            'admission_date': student.admission_date.strftime('%Y-%m-%d'),
            'class_ids': class_ids  # Trả về danh sách class_id
        })
    return jsonify({'error': 'Student not found'}), 404


@app.route('/update-student/<int:student_id>', methods=['POST'])
@login_required
def update_student(student_id):
    try:
        data = request.form  # Dùng FormData khi gửi từ frontend
        print(f"Received Data: {data}")  # Debug log

        # Kiểm tra sự tồn tại của học sinh
        student = Student.query.get(student_id)
        if not student:
            return jsonify({'success': False, 'error': 'Student not found'}), 404

        # Cập nhật thông tin sinh viên
        student.name = data.get('name', student.name)
        student.dateOfBirth = datetime.strptime(data.get('date_of_birth', student.dateOfBirth.strftime('%Y-%m-%d')),
                                                '%Y-%m-%d')
        student.gender = data.get('gender', student.gender)
        student.address = data.get('address', student.address)
        student.phoneNumber = data.get('phone_number', student.phoneNumber)
        student.email = data.get('email', student.email)
        student.admission_date = datetime.strptime(
            data.get('admission_date', student.admission_date.strftime('%Y-%m-%d')), '%Y-%m-%d')

        # Lưu thay đổi vào database
        db.session.commit()
        return jsonify({'success': True, 'message': 'Student updated successfully'})

    except Exception as e:
        print(f"Error: {e}")
        db.session.rollback()
        return jsonify({'success': False, 'error': 'Internal server error'}), 500


# Dinh tuyen trang danh sach lop
@app.route('/dslop', methods=['GET', 'POST'])
@login_required
def dslophoc():
    class_list = dao.get_class_list()
    return render_template('staff/dslophoc.html', class_list=class_list)


@app.route('/get_danh_sach_lop', methods=['GET'])
@login_required
def get_danh_sach_lop():
    danh_sach_lop = Class.query.all()
    return jsonify([{'id': lop.id, 'className': lop.className, 'quantity': lop.quantity} for lop in danh_sach_lop])


@app.route('/get_lop/<int:class_id>', methods=['GET'])
@login_required
def get_lop(class_id):
    lop = dao.get_class_by_id(class_id)
    return jsonify({'id': lop.id, 'className': lop.className, 'quantity': lop.quantity})


@app.route('/sua_lop/<int:class_id>', methods=['POST'])
@login_required
def sua_lop(class_id):
    lop = dao.get_class_by_id(class_id)

    if request.method == 'POST':
        ten_lop_moi = request.form['className']
        si_so_moi = int(request.form['quantity'])

        lop.className = ten_lop_moi
        lop.quantity = si_so_moi
        db.session.commit()

        return jsonify({'message': 'Thông tin lớp học đã được cập nhật thành công.'})


# Thêm lop hoc
@app.route('/add_class', methods=['POST'])
def add_class():
    data = request.get_json()
    class_name = data.get('className')
    quantity = data.get('quantity')
    grade = data.get('grade')  # Giá trị gửi từ frontend

    # Kiểm tra dữ liệu đầu vào
    if not class_name or not quantity or not grade:
        return jsonify({'error': 'Dữ liệu không hợp lệ'}), 400

    try:
        # Chuyển đổi grade sang Enum StudentGrade
        grade_enum = StudentGrade[grade]  # Sử dụng tên Enum, ví dụ: "GRADE_10TH"

        # Thêm lớp học vào cơ sở dữ liệu
        new_class = Class(className=class_name, quantity=quantity, grade=grade_enum)
        db.session.add(new_class)
        db.session.commit()

        return jsonify({'message': 'Lớp học được thêm thành công'}), 201
    except KeyError:
        return jsonify({'error': 'Mã khối không hợp lệ'}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500




@app.route('/xoa_lop/<int:class_id>', methods=['DELETE'])
@login_required
def xoa_lop(class_id):
    # Tìm lớp theo class_id
    lop = db.session.query(Class).filter(Class.id == class_id).first()

    if not lop:
        return jsonify({'error': 'Lớp không tồn tại.'}), 404

    # Xóa các bản ghi liên quan trong bảng StudentClass
    db.session.query(StudentClass).filter(StudentClass.class_id == class_id).delete()

    # Xóa các bản ghi liên quan trong bảng Teach
    db.session.query(Teach).filter(Teach.class_id == class_id).delete()

    # Xóa lớp
    db.session.delete(lop)
    db.session.commit()

    return jsonify({'message': 'Lớp học đã được xóa thành công.'})


# Dinh tuyen trang Teacher
@app.route('/teacher')
@login_required
def teacher():
    return render_template('teacher/teacher.html')


# Dinh tuyen trang nhap diem
@app.route('/nhapdiem')
@login_required
def input_score():
    class_list = dao.get_class_list()
    subject = dao.get_subject_list()
    semester = dao.get_semester()
    return render_template('teacher/nhapdiem.html', class_list=class_list, subject=subject, semester=semester)


# Dinh tuyen trang xuat diem
@app.route('/xuatdiem')
@login_required
def output_score():
    class_list = dao.get_class_list()
    subject = dao.get_subject_list()
    return render_template('teacher/xuatdiem.html', class_list=class_list, subject=subject)


@app.route('/save_scores', methods=['POST'])
@login_required
def save_scores():
    data = request.get_json()  # Nhận dữ liệu từ client
    scores = data.get('scores', [])  # Lấy danh sách điểm

    if not scores:
        return jsonify({'error': 'No scores provided'}), 400

    # Kiểm tra xem có subject_id trong mỗi điểm không
    for score in scores:
        if 'subject_id' not in score:
            return jsonify({'error': 'subject_id missing in score'}), 400

    try:
        # Lặp qua tất cả điểm và lưu vào cơ sở dữ liệu
        for score in scores:
            student_id = score['student_id']
            score_15p = score.get('score_15p')
            score_1tiet = score.get('score_1tiet')
            score_exam = score.get('score_exam')
            subject_id = score['subject_id']
            semester_id = score['semester_id']

            # Kiểm tra và chỉ lưu điểm 15 phút nếu không phải chuỗi rỗng hoặc None
            if score_15p not in (None, ''):  # Nếu điểm không phải None hoặc chuỗi rỗng
                score_detail_15p = ScoreDetail(
                    score=score_15p,
                    type=ScoreType.EXAM_15MINS,
                    student_id=student_id
                )
                db.session.add(score_detail_15p)
                db.session.commit()  # Lưu vào bảng ScoreDetail

                new_score_15p = Score(
                    score=score_15p,
                    type=ScoreType.EXAM_15MINS,
                    student_id=student_id,
                    scoreDetail_id=score_detail_15p.id,
                    semester_id=semester_id,
                    subject_id=subject_id
                )
                db.session.add(new_score_15p)

            # Kiểm tra và chỉ lưu điểm 1 tiết nếu không phải chuỗi rỗng hoặc None
            if score_1tiet not in (None, ''):
                score_detail_1tiet = ScoreDetail(
                    score=score_1tiet,
                    type=ScoreType.EXAM_45MINS,
                    student_id=student_id
                )
                db.session.add(score_detail_1tiet)
                db.session.commit()

                new_score_1tiet = Score(
                    score=score_1tiet,
                    type=ScoreType.EXAM_45MINS,
                    student_id=student_id,
                    scoreDetail_id=score_detail_1tiet.id,
                    semester_id=semester_id,
                    subject_id=subject_id
                )
                db.session.add(new_score_1tiet)

            # Kiểm tra và chỉ lưu điểm thi cuối kỳ nếu không phải chuỗi rỗng hoặc None
            if score_exam not in (None, ''):
                score_detail_exam = ScoreDetail(
                    score=score_exam,
                    type=ScoreType.EXAM_FINAL,
                    student_id=student_id
                )
                db.session.add(score_detail_exam)
                db.session.commit()

                new_score_exam = Score(
                    score=score_exam,
                    type=ScoreType.EXAM_FINAL,
                    student_id=student_id,
                    scoreDetail_id=score_detail_exam.id,
                    semester_id=semester_id,
                    subject_id=subject_id
                )
                db.session.add(new_score_exam)

        db.session.commit()  # Commit toàn bộ giao dịch

        return jsonify({'message': 'Scores saved successfully'}), 200

    except Exception as e:
        db.session.rollback()  # Rollback trong trường hợp có lỗi
        print(f"Error: {e}")
        return jsonify({'error': str(e)}), 500


# Dinh tuyen danh sach mon
@app.route('/dsmon')
@login_required
def dsmonhoc():
    subject = dao.get_subject_list()
    return render_template('teacher/dsmon.html', subject=subject)


# Hiện thị học sinh theo lớp
@app.route('/get_data', methods=['POST'])
@login_required
def get_data():
    # Get the selected class ID from the JSON request data
    request_data = request.get_json()
    selected_class_id = request_data.get('class_id')
    # Query the database for students in the selected class
    data = db.session.query(StudentClass, Student).join(Student).filter(
        StudentClass.class_id == selected_class_id).all()
    class_name = db.session.query(Class).filter(Class.id == selected_class_id).first().className
    # Format the data as a list of dictionaries
    formatted_data = {
        'students': [
            {
                'id': student.id,
                'name': student.name,
                'gender': str(student.gender.value),
                'dateOfBirth': student.dateOfBirth,  # Ensure this is the correct attribute name
                'address': student.address
            }
            for student_class, student in data
        ],
        'class_name': class_name  # You can replace this with actual class name if available
    }

    return jsonify(formatted_data)


@app.route('/get_datas', methods=['POST'])
@login_required
def get_datas():
    # Lấy dữ liệu từ request
    request_data = request.get_json()
    selected_class_id = request_data.get('class_id')
    subject_id = request_data.get('subject_id')

    # Lấy thông tin học sinh thuộc lớp đã chọn
    student_classes = db.session.query(StudentClass, Student).join(Student).filter(
        StudentClass.class_id == selected_class_id
    ).all()

    # Lấy tên lớp
    class_info = db.session.query(Class).filter(Class.id == selected_class_id).first()
    class_name = class_info.className if class_info else None

    # Lấy thông tin điểm số cho môn học đã chọn
    scores = Score.query.filter_by(subject_id=subject_id).all()

    # Tạo danh sách thông tin học sinh kèm điểm số
    student_data = []
    for student_class, student in student_classes:
        # Lọc điểm của từng học sinh
        student_scores = [score for score in scores if score.student_id == student.id]

        # Tìm các loại điểm theo kiểu điểm
        exam_15mins = next((s.score for s in student_scores if s.type == ScoreType.EXAM_15MINS), None)
        exam_45mins = next((s.score for s in student_scores if s.type == ScoreType.EXAM_45MINS), None)
        final_exam = next((s.score for s in student_scores if s.type == ScoreType.EXAM_FINAL), None)

        # Thêm thông tin học sinh vào danh sách
        student_data.append({
            'id': student.id,
            'name': student.name,
            'exam_15mins': exam_15mins,
            'exam_45mins': exam_45mins,
            'final_exam': final_exam,
            'semester_id': student.semester_id
        })

    # Debug log
    print(f'Class ID: {selected_class_id}, Subject ID: {subject_id}')
    print(f'Student Data: {student_data}')
    print(f'Scores: {scores}')

    # Trả về dữ liệu dưới dạng JSON
    return jsonify({
        'students': student_data,
        'class_name': class_name
    })


@app.route('/export_excel', methods=['POST'])
def export_excel():
    # Lấy class_id và subject_id từ form data
    class_id = request.form.get('class_id')
    subject_id = request.form.get('subject_id')

    print("Class ID:", class_id)
    print("Subject ID:", subject_id)

    # Truy vấn điểm từ cơ sở dữ liệu
    scores = (
        db.session.query(Score)
        .join(Student, Score.student_id == Student.id)
        .join(StudentClass, StudentClass.student_id == Student.id)
        .join(Subject, Score.subject_id == Subject.id)
        .filter(StudentClass.class_id == class_id, Score.subject_id == subject_id)
        .all()
    )

    if not scores:
        return "No data found", 404
    else:
        print(f"Found {len(scores)} scores.")

    # Tạo workbook và worksheet
    wb = Workbook()
    ws = wb.active

    # Thiết lập tiêu đề
    ws.merge_cells('A1:E1')
    ws['A1'] = 'BẢNG ĐIỂM MÔN HỌC'
    ws['A1'].alignment = Alignment(horizontal='center')

    # Truy vấn thông tin lớp học, môn học, năm học và học kỳ
    class_instance = Class.query.filter_by(id=class_id).first()
    subject_instance = Subject.query.filter_by(id=subject_id).first()
    semester = Semester.query.first()  # Hoặc query theo cách bạn lấy năm học và học kỳ

    # Thêm thông tin lớp học, môn học, năm học và học kỳ vào worksheet
    ws['A3'] = 'Lớp:'
    ws['B3'] = class_instance.className if class_instance else "Unknown Class"
    ws['A4'] = 'Môn:'
    ws['B4'] = subject_instance.subjectName if subject_instance else "Unknown Subject"
    ws['A5'] = 'Năm học:'
    ws['B5'] = semester.year if semester else "Unknown Year"
    ws['A6'] = 'Học kỳ:'
    ws['B6'] = str(semester.semester.value) if semester else "Unknown Semester"

    # Thêm tiêu đề các cột
    headers = ['STT', 'Họ tên', 'Điểm 15 phút', 'Điểm 1 tiết', 'Điểm thi']
    ws.append(headers)

    # Lặp qua các sinh viên và điểm của họ
    for idx, student in enumerate(set(score.student for score in scores), start=1):
        # Truy xuất các giá trị thực từ đối tượng sinh viên và điểm
        student_name = student.name if student else "Unknown"
        score_15mins = next((score.score for score in scores if score.type == ScoreType.EXAM_15MINS), "N/A")
        score_45mins = next((score.score for score in scores if score.type == ScoreType.EXAM_45MINS), "N/A")
        score_final = next((score.score for score in scores if score.type == ScoreType.EXAM_FINAL), "N/A")

        print(f"Student: {student_name}, 15 min: {score_15mins}, 45 min: {score_45mins}, Final: {score_final}")

        # Thêm dữ liệu vào bảng Excel
        ws.append([
            idx,
            student_name,
            score_15mins,
            score_45mins,
            score_final
        ])

    # Đặt đường dẫn lưu file
    file_path = os.path.join(os.getcwd(), 'student_scores.xlsx')

    # Lưu workbook
    wb.save(file_path)

    # Kiểm tra nếu file đã tồn tại và gửi file về phía client
    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True)
    else:
        return "File not found", 404


@app.route('/export_avg_scores', methods=['POST'])
def export_avg_scores():
    # Sử dụng session của Flask-SQLAlchemy
    session = db.session

    # Truy vấn điểm trung bình của sinh viên cho học kỳ 1 và học kỳ 2
    avg_scores_query = (
        session.query(
            Student.id.label(""),
            Student.name.label("student_name"),
            StudentClass.class_id,
            Class.className.label("class_name"),
            func.round(
                (
                    func.sum(
                        case(
                            (Score.type == ScoreType.EXAM_15MINS, Score.score),
                            (Score.type == ScoreType.EXAM_45MINS, Score.score),
                            (Score.type == ScoreType.EXAM_FINAL, Score.score),
                            else_=0
                        )
                    )
                ) / 3, 2
            ).label("average_score"),
            Semester.year.label("year"),
            Semester.semester.label("semester"),
        )
        .join(Score, Student.id == Score.student_id)
        .join(Semester, Score.semester_id == Semester.id)
        .join(StudentClass, Student.id == StudentClass.student_id)
        .join(Class, StudentClass.class_id == Class.id)
        .group_by(Student.id, Semester.year, Semester.semester, Class.id)
    )

    # Lấy dữ liệu từ truy vấn
    data = avg_scores_query.all()

    # Tạo workbook và worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Bảng Điểm Trung Bình"

    # Ghi tiêu đề
    ws.merge_cells('A1:E1')
    ws['A1'] = "BẢNG ĐIỂM TRUNG BÌNH HỌC KỲ"

    # Lấy năm học từ dữ liệu đầu tiên (giả sử tất cả học sinh cùng một năm học)
    year = data[0].year if data else "Unknown Year"
    ws['A2'] = f"Năm học: {year}"

    # Ghi header bảng
    headers = ["STT", "Họ tên", "Lớp", "Điểm TB HK1", "Điểm TB HK2"]
    ws.append(headers)

    # Ghi dữ liệu
    for idx, student in enumerate(data, start=1):
        # Điểm học kỳ 1 và học kỳ 2
        if student.semester == SemesterType.SEMESTER_1:
            hk1_score = round(student.average_score, 2) if student.average_score is not None else "N/A"
            hk2_score = "N/A"  # Điểm học kỳ 2 nếu có thể tính hoặc lấy từ một truy vấn khác
        else:
            hk1_score = "N/A"
            hk2_score = round(student.average_score, 2) if student.average_score is not None else "N/A"

        # Ghi vào bảng Excel
        ws.append([
            idx,
            student.student_name,
            student.class_name,  # Tên lớp
            hk1_score,  # Điểm học kỳ 1
            hk2_score  # Điểm học kỳ 2
        ])

    # Lưu vào buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Trả file về client
    return send_file(
        buffer,
        as_attachment=True,
        download_name="bang_diem_trung_binh.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


if __name__ == "__main__":
    from ManageApp.admin import *

    app.run(debug=False)
